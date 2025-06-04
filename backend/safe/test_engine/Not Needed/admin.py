#backend/test_engine/admin.py
from django.contrib import admin
from django.urls import path
from django.shortcuts import redirect, render
from django.utils.html import format_html
from django.utils.text import slugify
from django import forms
from django.contrib import messages
import csv, io, random, os, openpyxl, zipfile
from django.http import HttpResponse
from openpyxl.styles import Font
from decimal import Decimal
import re
import ast  # safer for parsing literal Python-style lists
import json


from .models import (
    Test, Question, Candidate, Response, ScoreReport,
    TestSectionConfig, TestQuestionSet, QuestionCategory, ScoreReport
)

# ---- CSV Upload Form ---- #

class CSVUploadForm(forms.Form):
    csv_file = forms.FileField()

# ---- Question Category Admin (with CSV import) ---- #

@admin.register(QuestionCategory)
class QuestionCategoryAdmin(admin.ModelAdmin):
    list_display = ['id', 'name', 'import_questions_link']
    search_fields = ['name']

    def import_questions_link(self, obj):
        return format_html(
            '<a href="/admin/test_engine/questioncategory/import-csv/">Import CSV</a>'
        )
    import_questions_link.short_description = 'Import Questions'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-csv/', self.admin_site.admin_view(self.import_csv_view), name='import-questions-by-category'),
        ]
        return custom_urls + urls

    def import_csv_view(self, request):
        if request.method == 'POST':
            category_id = request.POST.get('category_id')
            form = CSVUploadForm(request.POST, request.FILES)

            if form.is_valid() and category_id:
                csv_file = form.cleaned_data['csv_file']
                decoded_file = csv_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                category = QuestionCategory.objects.get(id=category_id)

                created, skipped, errors = 0, 0, []

                def clean_options(option_str):
                    try:
                        # Try JSON load directly
                        parsed = json.loads(option_str)
                        if isinstance(parsed, list):
                            return json.dumps(parsed)
                    except Exception:
                        pass

                    try:
                        # Try Python literal evaluation (handles single quotes)
                        parsed = ast.literal_eval(option_str)
                        if isinstance(parsed, list):
                            return json.dumps(parsed)
                    except Exception:
                        pass

                    raise ValueError(f"❌ Invalid options format: {option_str}")
                
                for i, row in enumerate(reader, start=2):  # start=2 (header is row 1)
                    try:
                        print(f"👉 Raw CSV input (row {i}):", row['options'])

                        options_clean = clean_options(row['options'])
                        Question.objects.create(
                            category=category,
                            text=row['text'],
                            difficulty=row['difficulty'],
                            question_type=row.get('question_type', 'MCQ'),
                            correct_answer=row['correct_answer'],
                            options=options_clean,
                            positive_marks=float(row.get('positive_marks', 1.0)),
                            negative_marks=float(row.get('negative_marks', 0.0))
                        )
                        created += 1
                    except Exception as e:
                        skipped += 1
                        errors.append(f"Row {i}: {e}")

                # Show message
                msg = f"✅ Imported {created} questions to category '{category.name}'."
                if skipped:
                    msg += f" ⚠️ Skipped {skipped} rows."
                    for err in errors[:5]:
                        msg += f"\n - {err}"
                    if skipped > 5:
                        msg += f"\n - ... and {skipped - 5} more."

                self.message_user(request, msg, level=messages.WARNING if skipped else messages.SUCCESS)
                return redirect('/admin/test_engine/question/')

        else:
            form = CSVUploadForm()

        categories = QuestionCategory.objects.all()
        return render(request, 'admin/import_questions_by_category.html', {'form': form, 'categories': categories})


# ---- Test Admin ---- #

@admin.register(Test)
class TestAdmin(admin.ModelAdmin):
    list_display = ['id', 'name', 'generate_button']

    list_display = [
        'id',
        'name',
        'total_duration_minutes',
        'total_section_time_display',
        'enforce_section_time',
        'show_section_time_guidance',
        'generate_button',
    ]

    def total_section_time_display(self, obj):
        return obj.total_section_time()

    total_section_time_display.short_description = "Sum of Section Durations"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:test_id>/generate-questions/', self.admin_site.admin_view(self.generate_questions), name='generate-questions'),
        ]
        return custom_urls + urls

    def generate_button(self, obj):
        return format_html(
            '<a class="button" href="{}">Generate Questions</a>',
            f'{obj.id}/generate-questions/'
        )
    generate_button.short_description = 'Actions'
    generate_button.allow_tags = True

    def generate_questions(self, request, test_id):
        test = Test.objects.get(pk=test_id)
        sections = test.sections.all()

        # Clear existing generated questions
        TestQuestionSet.objects.filter(test=test).delete()

        question_set = []
        order_counter = 0

        for section in sections:
            for difficulty, count in [
                ('easy', section.easy_questions),
                ('moderate', section.moderate_questions),
                ('hard', section.hard_questions),
            ]:
                questions = Question.objects.filter(
                    category=section.category,
                    difficulty=difficulty
                ).order_by('?')[:count]

                if questions.count() < count:
                    messages.warning(
                        request,
                        f"Not enough {difficulty} questions in category '{section.category.name}' "
                        f"for test '{test.name}'. Requested {count}, found {questions.count()}."
                    )

                for q in questions:
                    question_set.append(TestQuestionSet(
                        test=test,
                        question=q,
                        order=order_counter
                    ))
                    order_counter += 1

        TestQuestionSet.objects.bulk_create(question_set)
        messages.success(request, f"Generated {len(question_set)} questions for test: {test.name}")
        return redirect(f'/admin/test_engine/test/{test_id}/change/')


# ---- Other Admin Classes ---- #

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ['id', 'text', 'category', 'difficulty', 'question_type']
    search_fields = ['text']
    list_filter = ['category', 'difficulty']
    fields = [
        'text',
        'question_type',
        'options',
        'correct_answer',
        'category',
        'difficulty',
        'positive_marks',
        'negative_marks'
    ]

    def formfield_for_dbfield(self, db_field, **kwargs):
        formfield = super().formfield_for_dbfield(db_field, **kwargs)
        if db_field.name == 'options':
            formfield.help_text = 'Enter choices as a JSON list, e.g., ["A", "B", "C", "D"]'
        return formfield

    def save_model(self, request, obj, form, change):
        import json

        if isinstance(obj.options, str):
            try:
                # Validate as-is — do NOT auto-fix
                parsed = json.loads(obj.options)
                if not isinstance(parsed, list):
                    raise ValueError
            except Exception:
                raise ValueError(
                    f"⚠️ Invalid options format.\nMust be a valid JSON list of strings, like: [\"A\", \"B\", \"C\"]\n\nGot: {obj.options}"
                )
        super().save_model(request, obj, form, change)


from django.db.models import Count, Q
from django.template.response import TemplateResponse

@admin.register(TestSectionConfig)
class TestSectionConfigAdmin(admin.ModelAdmin):
    list_display = [
        'test', 'category',
        'easy_questions', 'moderate_questions', 'hard_questions',
        'section_duration_minutes', 'total_questions_display'
    ]


    def changelist_view(self, request, extra_context=None):
        # Summary logic
        from .models import QuestionCategory, Question
        summary = []
        categories = QuestionCategory.objects.all()

        for cat in categories:
            easy = Question.objects.filter(category=cat, difficulty='easy').count()
            moderate = Question.objects.filter(category=cat, difficulty='moderate').count()
            hard = Question.objects.filter(category=cat, difficulty='hard').count()
            summary.append((cat.name, easy, moderate, hard))

        extra_context = extra_context or {}
        extra_context['summary'] = summary
        return super().changelist_view(request, extra_context=extra_context)

    def total_questions_display(self, obj):
        return obj.total_questions()

    total_questions_display.short_description = "Total Questions"


@admin.register(TestQuestionSet)
class TestQuestionSetAdmin(admin.ModelAdmin):
    list_display = ['test', 'question', 'order']
    list_filter = ['test']
    search_fields = ['question__text']


@admin.register(Candidate)
class CandidateAdmin(admin.ModelAdmin):
    list_display = ['id', 'name', 'email']
    search_fields = ['email']

@admin.register(Response)
class ResponseAdmin(admin.ModelAdmin):
    list_display = ['id', 'candidate', 'question', 'answer']



# ---- Export Score Report ---- #

@admin.action(description="Export selected scores to Excel")
def export_scores_to_excel(modeladmin, request, queryset):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font
    from .models import Response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"

    # First, collect all unique categories used in the test set
    all_categories = set()
    for report in queryset:
        cat_ids = report.test.sections.values_list("category__name", flat=True)
        all_categories.update(cat_ids)

    sorted_categories = sorted(all_categories)  # So columns remain consistent

    # Build header row
    base_headers = [
        "Candidate Name", "Email", "Phone", "Test",
        "Score", "Max Possible", "Percentage",
        "Total Correct", "Total Wrong", "Unattempted"
    ]

    category_headers = []
    for cat in sorted_categories:
        prefix = f"{cat}"
        category_headers.extend([
            f"{prefix} _ Score",
            f"{prefix} _ Max Possible",
            f"{prefix} _ Percentage",
            f"{prefix} _ Total Correct",
            f"{prefix} _ Total Wrong",
            f"{prefix} _ Unattempted"
        ])

    headers = base_headers + category_headers + ["Created At"]
    ws.append(headers)

    # Bold header row
    bold_font = Font(bold=True)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = bold_font

    # For each report
    for report in queryset.select_related("candidate", "test"):
        base_data = [
            report.candidate.name,
            report.candidate.email,
            report.candidate.phone,
            report.test.name,
            float(report.score),
            float(report.total_positive + report.total_negative),
            round((report.score / (report.total_positive + report.total_negative) * 100) if (report.total_positive + report.total_negative) else 0, 2),
            report.total_correct,
            report.total_wrong,
            report.total_unattempted,
        ]

        # Prepare per-category data
        responses = Response.objects.filter(
            candidate=report.candidate,
            question__category__in=report.test.sections.values_list('category', flat=True)
        ).select_related('question__category')

        by_category = {}

        for r in responses:
            cat = r.question.category.name
            if cat not in by_category:
                by_category[cat] = {
                    'score': 0,
                    'max': 0,
                    'correct': 0,
                    'wrong': 0,
                    'unattempted': 0
                }

            submitted = (r.answer or "").strip().lower()
            correct = (r.question.correct_answer or "").strip().lower()

            by_category[cat]['max'] += r.question.positive_marks

            if not submitted:
                by_category[cat]['unattempted'] += 1
            elif submitted == correct:
                by_category[cat]['correct'] += 1
                by_category[cat]['score'] += r.question.positive_marks
            else:
                by_category[cat]['wrong'] += 1
                by_category[cat]['score'] -= r.question.negative_marks
                by_category[cat]['max'] += r.question.negative_marks

        # Add values in sorted order
        category_data = []
        for cat in sorted_categories:
            cat_data = by_category.get(cat, {
                'score': 0,
                'max': 0,
                'correct': 0,
                'wrong': 0,
                'unattempted': 0
            })

            percent = (cat_data['score'] / cat_data['max'] * 100) if cat_data['max'] else 0
            category_data.extend([
                round(cat_data['score'], 2),
                round(cat_data['max'], 2),
                round(percent, 2),
                cat_data['correct'],
                cat_data['wrong'],
                cat_data['unattempted']
            ])

        ws.append(base_data + category_data + [report.created_at.strftime("%Y-%m-%d %H:%M:%S")])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=score_export_detailed.xlsx"
    wb.save(response)
    return response


# ---- Export Evaluated Answers ---- #
@admin.action(description="Export evaluated answer sheets (Excel)")
def export_evaluated_answers(modeladmin, request, queryset):
    import io
    import zipfile
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font
    from django.utils.text import slugify
    from django.utils import timezone
    from datetime import timedelta
    from .models import Response, Question, TestQuestionSet, TestSectionConfig, CandidateTestSession

    files = []

    for report in queryset.select_related('candidate', 'test'):
        wb = openpyxl.Workbook()
        main_ws = wb.active
        main_ws.title = "Answer Sheet"
        summary_ws = wb.create_sheet(title="Category Summary")

        # === SHEET 1 HEADERS ===
        headers = [
            "#", "Category", "Difficulty", "Question",
            "Your Answer (Choice)", "Your Answer (Raw)",
            "Correct Answer (Choice)", "Correct Answer (Raw)",
            "Status", "+Marks", "-Marks", "Score Awarded", "Time Taken (s)", "Answered At", "Within Time?"
        ]
        main_ws.append(headers)
        for cell in main_ws[1]:
            cell.font = Font(bold=True)

        tq_ids = TestQuestionSet.objects.filter(test=report.test).values_list('question_id', flat=True)
        questions = Question.objects.filter(id__in=tq_ids).select_related('category')

        responses = {
            r.question_id: r for r in Response.objects.filter(
                candidate=report.candidate,
                question_id__in=[q.id for q in questions]
            )
        }

        # Pull section time info
        session = CandidateTestSession.objects.filter(candidate=report.candidate, test=report.test).first()
        section_map = {
            s.category_id: (s.section_duration_minutes, session.section_started_at if session and session.section_started_at else None)
            for s in TestSectionConfig.objects.filter(test=report.test)
        }

        category_summary = {}

        def get_choice_letter(val, options):
            try:
                return chr(options.index(val) + ord('A'))
            except:
                return ""

        for i, q in enumerate(questions, start=1):
            r = responses.get(q.id)
            raw_submitted = (r.answer or "").strip() if r else ""
            raw_correct = (q.correct_answer or "").strip()

            try:
                opts = eval(q.options)
            except:
                opts = []

            def label_all(raw):
                return ",".join(get_choice_letter(val, opts) for val in raw.split(",") if get_choice_letter(val, opts))

            submitted_choice = label_all(raw_submitted)
            correct_choice = label_all(raw_correct)

            status = "Unattempted"
            awarded = 0
            if raw_submitted:
                if sorted(raw_submitted.split(",")) == sorted(raw_correct.split(",")):
                    status = "Correct"
                    awarded = q.positive_marks
                else:
                    status = "Wrong"
                    awarded = -q.negative_marks

            cat = q.category.name
            if cat not in category_summary:
                category_summary[cat] = {
                    "total_qs": 0, "correct": 0, "wrong": 0,
                    "unattempted": 0, "positive": 0, "negative": 0,
                    "max_marks": 0, "section_start": None, "duration": None
                }
            stats = category_summary[cat]
            stats["total_qs"] += 1
            stats["max_marks"] += q.positive_marks
            if status == "Correct":
                stats["correct"] += 1
                stats["positive"] += q.positive_marks
            elif status == "Wrong":
                stats["wrong"] += 1
                stats["negative"] += q.negative_marks
            else:
                stats["unattempted"] += 1

            # Determine if answered within time
            within_time = ""
            deadline = None
            answered_at_str = ""
            if r and hasattr(r, 'answered_at') and r.answered_at:
                answered_at_str = r.answered_at.strftime("%Y-%m-%d %H:%M:%S")
                sec = section_map.get(q.category_id)
                if sec and sec[1]:
                    try:
                        deadline = sec[1] + timedelta(minutes=sec[0])
                        within_time = "Yes" if r.answered_at <= deadline else "No"
                    except:
                        within_time = ""

            main_ws.append([
                i,
                cat,
                q.difficulty,
                q.text[:200].replace("\n", " "),
                submitted_choice,
                raw_submitted,
                correct_choice,
                raw_correct,
                status,
                q.positive_marks,
                q.negative_marks,
                awarded,
                r.time_spent if r else "",
                answered_at_str,
                within_time
            ])

        summary_headers = [
            "Category", "Total", "Correct", "Wrong", "Unattempted",
            "Total +Marks", "Total -Marks", "Net Score", "Max Score", "Percentage",
            "Section Start", "Allotted Duration (min)", "Should End Time"
        ]
        summary_ws.append(summary_headers)
        for cell in summary_ws[1]:
            cell.font = Font(bold=True)

        # Fill category summary
        grand = {
            "total_qs": 0, "correct": 0, "wrong": 0, "unattempted": 0,
            "positive": 0, "negative": 0, "max_marks": 0
        }

        for cat, s in category_summary.items():
            net_score = s["positive"] - s["negative"]
            percent = ((s["positive"] - s["negative"]) / s["max_marks"] * 100) if s["max_marks"] else 0

            section = TestSectionConfig.objects.filter(test=report.test, category__name=cat).first()
            duration = section.section_duration_minutes if section else None
            section_start = session.section_started_at.strftime("%Y-%m-%d %H:%M:%S") if session and session.section_started_at else ""
            end_time = (session.section_started_at + timedelta(minutes=duration)).strftime("%Y-%m-%d %H:%M:%S") if session and session.section_started_at and duration else ""

            summary_ws.append([
                cat, s["total_qs"], s["correct"], s["wrong"], s["unattempted"],
                s["positive"], s["negative"], net_score, s["max_marks"], round(percent, 2),
                section_start, duration if duration is not None else "", end_time
            ])

            # Add to grand total
            for key in grand:
                grand[key] += s[key]

        # Append grand total row
        grand_net = grand["positive"] - grand["negative"]
        grand_percent = ((grand["positive"] - grand["negative"]) / grand["max_marks"] * 100) if grand["max_marks"] else 0
        summary_ws.append([
            "Grand Total", grand["total_qs"], grand["correct"], grand["wrong"], grand["unattempted"],
            grand["positive"], grand["negative"], grand_net, grand["max_marks"], round(grand_percent, 2),
            "", "", ""
        ])

        memfile = io.BytesIO()
        wb.save(memfile)
        memfile.seek(0)
        fname = f"{slugify(report.test.name)}_{slugify(report.candidate.name)}.xlsx"
        files.append((fname, memfile.read()))

    if len(files) == 1:
        name, content = files[0]
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={name}"
        return response
    else:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zf:
            for name, content in files:
                zf.writestr(name, content)
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type="application/zip")
        response["Content-Disposition"] = "attachment; filename=evaluated_answer_sheets.zip"
        return response







# ---- Recalculate Score ---- #


from .models import ScoreReport
from test_engine.utils.scoring import calculate_score_for_candidate

@admin.action(description="Recalculate score for selected reports")
def recalculate_scores(modeladmin, request, queryset):
    for report in queryset.select_related('candidate', 'test'):
        calculate_score_for_candidate(report.test, report.candidate)
    modeladmin.message_user(request, f"Recalculated score for {queryset.count()} report(s).")

@admin.register(ScoreReport)
class ScoreReportAdmin(admin.ModelAdmin):
    list_display = [
        'candidate', 'test', 'score',
        'max_possible', 'percentage',
        'total_correct', 'total_wrong', 'total_unattempted',
        'created_at'
    ]
    readonly_fields = ['score_breakdown_by_category']
    actions = [
        export_scores_to_excel,
        export_evaluated_answers,
        recalculate_scores
    ]

    def max_possible(self, obj):
        from .models import TestQuestionSet
        question_ids = TestQuestionSet.objects.filter(test=obj.test).values_list('question_id', flat=True)
        from test_engine.models import Question
        questions = Question.objects.filter(id__in=question_ids)
        total_positive = sum(q.positive_marks for q in questions)
        return round(total_positive, 2)

    max_possible.short_description = 'Max Possible'


    def percentage(self, obj):
        from .models import TestQuestionSet, Question

        question_ids = TestQuestionSet.objects.filter(test=obj.test).values_list('question_id', flat=True)
        questions = Question.objects.filter(id__in=question_ids)

        max_score = sum(Decimal(str(q.positive_marks)) for q in questions)

        return round((obj.score / max_score * 100), 2) if max_score > 0 else 0.0

    percentage.short_description = 'Percentage'

    def score_breakdown_by_category(self, obj):
        # Optional: show category-wise summary (can reuse earlier logic)
        return ""
